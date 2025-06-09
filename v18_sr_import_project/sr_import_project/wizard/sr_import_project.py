# -*- coding: utf-8 -*-
##############################################################################
#
#    OpenERP, Open Source Management Solution
#    Copyright (C) Sitaram Solutions (<https://sitaramsolutions.in/>).
#
#    For Module Support : info@sitaramsolutions.in  or Skype : contact.hiren1188
#
##############################################################################

import csv
import base64
import tempfile
from datetime import datetime
from odoo import models, fields, _
from openpyxl import load_workbook
from odoo.exceptions import UserError


class srImportProject(models.TransientModel):
    _name = 'import.project'
    _description = 'Import Project'

    file = fields.Binary('File')
    import_option = fields.Selection([('csv', 'CSV File'), ('xls', 'XLS File')], string='Select', default='csv')

    def import_project(self):
        if not self.file:
            raise UserError(_("Please upload a file to import."))

        data = base64.b64decode(self.file)
        if self.import_option == 'csv':
            self._import_csv(self.file)
        elif self.import_option == 'xls':
            self._import_xls(data)

    def _import_csv(self, data):
        try:
            if isinstance(data, bytes):
                file_content = base64.b64decode(data).decode("utf-8")
            else:
                file_content = data.decode("utf-8")

            csv_reader = csv.reader(file_content.splitlines(), delimiter=',')
            headers = next(csv_reader)

            for row in csv_reader:
                row_data = dict(zip(headers, row))

                self._create_project(row_data)
        except Exception as e:
            raise UserError(_("Error processing CSV file: %s") % str(e))

    def _import_xls(self, data):
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                tmp.write(data)
                tmp.flush()
                workbook = load_workbook(tmp.name, data_only=True)

            sheet = workbook.active
            headers = [cell.value for cell in sheet[1]]

            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_data = dict(zip(headers, row))

                self._create_project(row_data)
        except Exception as e:
            raise UserError(_("Error processing XLSX file: %s") % str(e))

    def _create_project(self, values):
        user_name = values.get('Project Manager')
        user_id = False
        if user_name:
            user_id = self.env['res.users'].search([('name', '=', user_name)], limit=1).id
            if not user_id and user_name:
                raise UserError(_("User '%s' not found") % user_name)

        customer_name = values.get('Customer')
        customer_id = False
        if customer_name:
            customer_id = self.env['res.partner'].search([('name', '=', customer_name)], limit=1).id
            if not customer_id and customer_name:
                raise UserError(_("Customer '%s' not found") % customer_name)

        tag_ids = []
        if values.get('Tags') and values['Tags']:
            tag_names = str(values.get('Tags', '')).split(',')
            for tag_name in tag_names:
                tag = self.env['project.tags'].search([('name', '=', tag_name.strip())], limit=1)
                if not tag:
                    raise UserError(_("Tags '%s' not found") % tag_name)
                tag_ids.append(tag.id)

        try:
            start_date = self._parse_date(values.get('Start Date'))
            end_date = self._parse_date(values.get('End Date'))
        except ValueError as e:
            raise UserError(_("Invalid date format: %s") % str(e))

        if not values.get('Name'):
            raise UserError(_("Name must be required.!"))
        project_vals = {
            'name': values.get('Name'),
            'partner_id': customer_id if customer_id else False,
            'user_id': user_id if user_id else False,
            'date_start': start_date,
            'date': end_date,
            'tag_ids': [(6, 0, tag_ids)],
            'description': values.get('Description', ''),
        }
        project = self.env['project.project'].create(project_vals)

        task_name = values.get('Task Name')
        if task_name:
            self.env['project.task'].create({
                'name': task_name,
                'project_id': project.id,
                'description': values.get('Description', ''),
            })

    def _parse_date(self, date_value):
        if not date_value:
            return None

        if isinstance(date_value, datetime):
            return date_value.date()
        try:
            return datetime.strptime(date_value, '%Y-%m-%d').date()
        except ValueError:
            raise ValueError(f"Date '{date_value}' is in an invalid format. Expected YYYY-MM-DD.")
