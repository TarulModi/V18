# -*- coding: utf-8 -*-
##############################################################################
#
#    OpenERP, Open Source Management Solution
#    Copyright (C) Sitaram Solutions (<https://sitaramsolutions.in/>).
#
#    For Module Support : info@sitaramsolutions.in  or Skype : contact.hiren1188
#
##############################################################################

import io
import csv
import xlrd
import base64
import logging
import tempfile
import binascii
from datetime import datetime
from odoo import models, fields, _
from openpyxl import load_workbook
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

class srImportStockPicking(models.TransientModel):
    _name = 'import.stock.picking'
    _description = 'Import Stock Picking'

    file = fields.Binary('File')
    import_option = fields.Selection([('csv', 'CSV File'), ('xls', 'XLS File')], string='Select', default='csv')
    picking_stage_selection = fields.Selection([('draft', 'Draft'), ('confirm', 'Confirm picking automatically when import')], string='Picking Stage', default='draft')
    sequence_option = fields.Selection([('custom', 'Use sequence from Excel/CSV'), ('default', 'Use Default Sequence')], string='Sequnce option', default='default')
    import_product_by = fields.Selection([('name', 'Name'), ('code', 'Code'), ('barcode', 'Barcode')], string='Import Product By', default='name')
    picking_type = fields.Selection([('incoming', 'Receipts'), ('outgoing', 'Deliveries'), ('internal', 'Internal')], string='Picking Type', default='incoming')
    source_location_id = fields.Many2one('stock.location', string="Source Location")
    destination_location_id = fields.Many2one('stock.location', string="Destination Location")

    def import_stock_picking(self):
        if not self.file:
            raise UserError(_("Please upload a file to import."))

        data = base64.b64decode(self.file)
        if self.import_option == 'csv':
            self._import_csv(self.file)
        elif self.import_option == 'xls':
            self._import_xls(data)

    # def _import_csv(self, data):
    #     try:
    #         file_content = data.decode("utf-8").splitlines()
    #         reader = csv.DictReader(file_content)
    #
    #         for row in reader:
    #             self._create_picking(row)
    #
    #     except Exception as e:
    #         raise UserError(_("Error processing CSV file: %s") % str(e))

    # def _import_xls(self, data):
    #     try:
    #         workbook = xlrd.open_workbook(file_contents=data)
    #         sheet = workbook.sheet_by_index(0)
    #
    #         headers = [sheet.cell_value(0, col) for col in range(sheet.ncols)]
    #
    #         for row_idx in range(1, sheet.nrows):
    #             row_data = {headers[col_idx]: sheet.cell_value(row_idx, col_idx) for col_idx in range(sheet.ncols)}
    #             self._create_picking(row_data)
    #
    #     except Exception as e:
    #         raise UserError(_("Error processing XLS file: %s") % str(e))

    # def _import_xls(self, data):
    #     try:
    #         with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
    #             tmp.write(data)
    #             tmp.flush()
    #             workbook = load_workbook(tmp.name, data_only=True)
    #
    #         sheet = workbook.active  # Get the first sheet
    #         headers = [cell.value for cell in sheet[1]]  # Read the first row as headers
    #
    #         for row in sheet.iter_rows(min_row=2, values_only=True):
    #             row_data = dict(zip(headers, row))  # Convert row to dictionary
    #             self._create_picking(row_data)
    #
    #     except Exception as e:
    #         raise UserError(_("Error processing XLSX file: %s") % str(e))

    # def _import_xls(self, data):
    #     try:
    #         with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
    #             tmp.write(data)
    #             tmp.flush()
    #             workbook = load_workbook(tmp.name, data_only=True)
    #
    #         sheet = workbook.active
    #         headers = [cell.value for cell in sheet[1]]  # Read first row as headers
    #
    #         _logger.info("Detected Headers: %s", headers)  # Log headers
    #
    #         print("========saaaaaaaaaaaaaaaaaaaaa=============",list(sheet.iter_rows(min_row=2, values_only=True)))
    #         for row in sheet.iter_rows(min_row=2, values_only=True):
    #             print("=============wwwwwwwwwwwwwwwwwwwwwwwwww===========",row)
    #             row_data = dict(zip(headers, row))  # Convert row to dictionary
    #             print("=============wwwwwwwwwwwwwwwwwwwwwwwwww======row_data=====",row_data)
    #
    #             _logger.info("Row Data: %s", row_data)  # Log row data
    #             print("========-----row_data---roe-------=============",row_data)
    #             self._create_picking(row_data)
    #
    #     except Exception as e:
    #         raise UserError(_("Error processing XLSX file: %s") % str(e))

    # def _create_picking(self, row):
    #     Picking = self.env['stock.picking']
    #     Move = self.env['stock.move']
    #
    #     print("----_create_picking-------------row---------row------",row)
    #     print("----_create_picking-------------row---------row---name---",row.get('Name'))
    #     print("----_create_picking-------------row---------row---Date---",row.get('Date'))
    #     print("----_create_picking-------------row---------row---Prcie---",row.get('Price'))
    #
    #     product = self._find_product(row)
    #     origin_value = f"Import Record {row.get('Name')}"
    #
    #     picking_vals = {
    #         'picking_type_id': self._get_picking_type(),
    #         'scheduled_date': row.get('Date'),
    #         'origin': origin_value,
    #         'state': 'draft' if self.picking_stage_selection == 'draft' else 'confirmed',
    #     }
    #
    #     # Conditional updates based on picking type
    #     if self.sequence_option == 'custom':
    #         if self.picking_type == 'incoming':
    #             picking_vals.update({
    #                 'name': row.get('Name'),
    #                 'location_dest_id': self.destination_location_id.id,
    #             })
    #         elif self.picking_type == 'outgoing':
    #             picking_vals.update({
    #                 'name': row.get('Name'),
    #                 'location_id': self.source_location_id.id,
    #             })
    #         else:
    #             picking_vals.update({
    #                 'name': row.get('Name'),
    #                 'location_id': self.source_location_id.id,
    #                 'location_dest_id': self.destination_location_id.id,
    #             })
    #     else:
    #         if self.picking_type == 'incoming':
    #             picking_vals.update({
    #                 'location_dest_id': self.destination_location_id.id,
    #             })
    #         elif self.picking_type == 'outgoing':
    #             picking_vals.update({
    #                 'location_id': self.source_location_id.id,
    #             })
    #         else:
    #             picking_vals.update({
    #                 'location_id': self.source_location_id.id,
    #                 'location_dest_id': self.destination_location_id.id,
    #             })
    #
    #     print("------picking_vals------picking-----", picking_vals)
    #     picking = Picking.create(picking_vals)
    #     print("------picking------picking-----",picking)
    #     print("------row.get('quantity', 1)------row.get('quantity', 1)-----",row.get('Quantity', 1))
    #
    #     move_vals = {
    #         'picking_id': picking.id,
    #         'product_id': product.id,
    #         'name': product.name,
    #         'product_uom_qty': float(row.get('Quantity', 1)),
    #         'product_uom': product.uom_id.id,
    #         'location_id': picking.location_id.id,
    #         'location_dest_id': picking.location_dest_id.id,
    #         'price_unit': float(row.get('Price', 0))
    #     }
    #     print("------move_vals------picking-----", move_vals)
    #     Move.create(move_vals)
    #
    #     if self.picking_stage_selection == 'confirm':
    #         picking.action_confirm()

    # def _find_product(self, row):
    #     Product = self.env['product.product']
    #     search_field = self.import_product_by
    #
    #     if self.import_product_by == 'code':
    #         search_value = 'default_code'
    #     elif self.import_product_by == 'barcode':
    #         search_value = 'barcode'
    #     else:
    #         search_value = 'name'
    #
    #     if not search_value:
    #         raise UserError(_("Missing product %s value in file.") % search_field)
    #
    #     product = Product.search([(search_field, '=', search_value)], limit=1)
    #     if not product:
    #         raise UserError(_("Product with %s '%s' not found.") % (search_field, search_value))
    #
    #     return product

    # def _import_xls(self, data):
    #     try:
    #         with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
    #             tmp.write(data)
    #             tmp.flush()
    #             workbook = load_workbook(tmp.name, data_only=True)
    #
    #         sheet = workbook.active
    #         # headers = [cell.value for cell in sheet[1]]  # Read first row as headers
    #         headers = [cell.value.strip().lower() for cell in sheet[1]]
    #
    #         _logger.info("Detected Headers: %s", headers)
    #
    #         pickings_data = {}  # Dictionary to store grouped picking data
    #
    #         # for row in sheet.iter_rows(min_row=2, values_only=True):
    #         #     row_data = dict(zip(headers, row))
    #         #
    #         #     picking_key = row_data.get('Source Document') or row_data.get(
    #         #         'Name')  # Group by 'Source Document' or 'Name'
    #         #
    #         #     if not picking_key:
    #         #         raise UserError(_("Missing 'Source Document' or 'Name' value in file."))
    #
    #         for row in sheet.iter_rows(min_row=2, values_only=True):
    #             print("-----row------------",row)
    #             row_data = dict(zip(headers, row))
    #             print("-----row_data------------",row_data)
    #
    #             # # Use lowercase keys to fetch values
    #             picking_key = row_data.get('source document') or row_data.get('name')
    #             print("-----picking_key------------",picking_key)
    #
    #             # if not picking_key:
    #             #     raise UserError(
    #             #         _("Missing 'Source Document' or 'Name' value in file. Available columns: %s") % row_data.keys())
    #
    #             if picking_key not in pickings_data:
    #                 pickings_data[picking_key] = {
    #                     'picking_vals': {
    #                         'picking_type_id': self._get_picking_type(),
    #                         'scheduled_date': row_data.get('Date'),
    #                         'origin': picking_key,
    #                         'state': 'draft' if self.picking_stage_selection == 'draft' else 'confirmed',
    #                     },
    #                     'products': []
    #                 }
    #
    #             pickings_data[picking_key]['products'].append(row_data)
    #
    #         for picking_key, data in pickings_data.items():
    #             self._create_picking(data['picking_vals'], data['products'])
    #
    #     except Exception as e:
    #         raise UserError(_("Error processing XLSX file: %s") % str(e))


    # def _find_product(self, row):
    #     """ Find product based on user selection (name, code, barcode) """
    #     Product = self.env['product.product']
    #
    #     search_field = self.import_product_by  # 'name', 'code', or 'barcode'
    #     search_value = row.get(search_field)  # Get actual value from the file
    #
    #     if not search_value:
    #         raise UserError(_("Missing product %s value in file.") % search_field)
    #
    #     product = Product.search([(search_field, '=', search_value)], limit=1)
    #
    #     if not product:
    #         raise UserError(_("Product with %s '%s' not found.") % (search_field, search_value))
    #
    #     return product

    # def _create_picking(self, picking_vals, product_rows):
    #     Picking = self.env['stock.picking']
    #     Move = self.env['stock.move']
    #
    #     picking = Picking.create(picking_vals)
    #     print("========picking=========",picking)
    #     _logger.info("Created Picking: %s", picking)
    #
    #     for row in product_rows:
    #         product = self._find_product(row)
    #
    #         move_vals = {
    #             'picking_id': picking.id,
    #             'product_id': product.id,
    #             'name': product.name,
    #             'product_uom_qty': float(row.get('Quantity', 1)),
    #             'product_uom': product.uom_id.id,
    #             'location_id': picking.location_id.id,
    #             'location_dest_id': picking.location_dest_id.id,
    #             'price_unit': float(row.get('Price', 0)),
    #         }
    #
    #         Move.create(move_vals)
    #         _logger.info("Added Product: %s (Qty: %s)", product.name, row.get('Quantity', 1))
    #
    #     if self.picking_stage_selection == 'confirm':
    #         picking.action_confirm()

    def _import_csv(self, data):
        """Method to import stock picking and products from a CSV file"""
        try:
            # Check if data is already a decoded string
            if isinstance(data, bytes):
                file_content = base64.b64decode(data).decode("utf-8")
            else:
                file_content = data.decode("utf-8")

            csv_reader = csv.reader(file_content.splitlines(), delimiter=',')

            headers = next(csv_reader)  # Read the first row as headers
            _logger.info("Detected Headers: %s", headers)

            last_name = None
            pickings = {}

            for row in csv_reader:
                row_data = dict(zip(headers, row))
                _logger.info("Row Data: %s", row_data)

                # If 'Name' is missing, use last known value
                name = row_data.get('Name') or last_name
                source_document = row_data.get('Source Document')
                date = row_data.get('Date') or row_data.get('date')
                partner = row_data.get('Contact')

                if not name:
                    raise UserError(_("Missing 'Name' value in file. Unable to proceed."))

                last_name = name  # Save last known order name

                # Find or create picking order
                if name not in pickings:
                    pickings[name] = self._create_picking(name, source_document, date, partner)

                # Add products (stock moves) to the picking
                self._create_move(pickings[name], row_data)

        except Exception as e:
            raise UserError(_("Error processing CSV file: %s") % str(e))

    def _import_xls(self, data):
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                tmp.write(data)
                tmp.flush()
                workbook = load_workbook(tmp.name, data_only=True)

            sheet = workbook.active
            headers = [cell.value for cell in sheet[1]]  # Read first row as headers

            _logger.info("Detected Headers: %s", headers)

            last_name = None
            # last_source_document = None
            pickings = {}

            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_data = dict(zip(headers, row))
                _logger.info("Row Data: %s", row_data)

                name = row_data.get('Name') or last_name
                source_document = row_data.get('Source Document')
                date = row_data.get('Date') or row_data.get('date')
                partner = row_data.get('Contact')

                if not name:
                    raise UserError(_("Missing 'Name' value in file. Unable to proceed."))

                # Update last known values
                last_name = name
                # last_source_document = source_document

                # Find or create picking
                if name not in pickings:
                    pickings[name] = self._create_picking(name, source_document, date, partner)

                # Add move lines (products) to the picking
                self._create_move(pickings[name], row_data)

        except Exception as e:
            raise UserError(_("Error processing XLSX file: %s") % str(e))

    def _create_picking(self, name, source_document, date, partner):
        """Create a new picking order"""
        Picking = self.env['stock.picking']
        print("================ddddddddddddddd===========",date)
        print("================ppppppppppppppppppppp===========",partner)

        # if date:
        #     scheduled_date = str(date) if isinstance(date, str) else date.strftime('%Y-%m-%d')
        # else:
        #     scheduled_date = False

        # scheduled_date = False
        # if date:
        #     try:
        #         # Detect various formats and convert to standard YYYY-MM-DD
        #         scheduled_date = datetime.strptime(date, '%d/%m/%y').strftime('%Y-%m-%d')
        #     except ValueError:
        #         try:
        #             scheduled_date = datetime.strptime(date, '%d/%m/%Y').strftime('%Y-%m-%d')
        #         except ValueError:
        #             print(f"Date format not recognized: {date}")
        #             scheduled_date = False  # Keep it False if parsing fails

        # print("================scheduled_date===========", scheduled_date)

        picking_vals = {
            'picking_type_id': self._get_picking_type(),
            # 'scheduled_date': scheduled_date,
            'partner_id' : self._find_partner(partner),
            'origin': source_document or f"Import Order {name}",
            'state': 'draft' if self.picking_stage_selection == 'draft' else 'confirmed',
        }
        if self.sequence_option == 'custom':
            if self.picking_type == 'incoming':
                picking_vals.update({
                    'name': name,
                    'location_dest_id': self.destination_location_id.id,
                })
            elif self.picking_type == 'outgoing':
                picking_vals.update({
                    'name': name,
                    'location_id': self.source_location_id.id,
                })
            else:
                picking_vals.update({
                    'name': name,
                    'location_id': self.source_location_id.id,
                    'location_dest_id': self.destination_location_id.id,
                })
        else:
            if self.picking_type == 'incoming':
                picking_vals.update({
                    'location_dest_id': self.destination_location_id.id,
                })
            elif self.picking_type == 'outgoing':
                picking_vals.update({
                    'location_id': self.source_location_id.id,
                })
            else:
                picking_vals.update({
                    'location_id': self.source_location_id.id,
                    'location_dest_id': self.destination_location_id.id,
                })
        print("-----picking_vals----picking_vals--------",picking_vals)
        picking = Picking.create(picking_vals)
        print("-----picking------------",picking)
        return picking

    def _find_partner(self, customer):
        if not customer:
            return False
        partner_id = self.env['res.partner'].search([('name', '=', customer)])
        if partner_id:
            return partner_id.id

    def _create_move(self, picking, row):
        """Create move lines (products) for the given picking"""
        Move = self.env['stock.move']
        product = self._find_product(row)

        move_vals = {
            'picking_id': picking.id,
            'product_id': product.id,
            'name': product.name,
            'product_uom_qty': float(row.get('Quantity', 1)),
            'product_uom': product.uom_id.id,
            'location_id': picking.location_id.id,
            'location_dest_id': picking.location_dest_id.id,
            'price_unit': float(row.get('Price', 0)),
        }
        Move.create(move_vals)

    def _find_product(self, row):
        print("----_find_product--------row-----",row)
        Product = self.env['product.product']
        search_field = self.import_product_by  # 'name', 'code', or 'barcode'
        print("----_find_product--------search_field-----",search_field)

        # Convert column names to lowercase for better matching
        row = {key.lower(): value for key, value in row.items()}
        print("----_find_product--------row--2222---",row)

        search_value = row.get('product name') or row.get('Product Name')
        print("----_find_product--------search_value-----",search_value)

        if not search_value:
            raise UserError(_("Missing product %s value in file. Available columns: %s") % (search_field, row.keys()))

        product = Product.search([(search_field, '=', search_value)], limit=1)
        print("----_find_product--------product-----",product)

        if not product:
            raise UserError(_("Product with %s '%s' not found. Check product records.") % (search_field, search_value))

        return product

    def _get_picking_type(self):
        picking_type = self.picking_type
        print("----picking_type-----_get_picking_type------",picking_type)
        if picking_type == 'internal':
            picking_type_id = self.env['stock.picking.type'].search([('code', '=', picking_type), ('sequence_code', '=', 'INT')], limit=1)
            print("----picking_type_id---_get_picking_type--------",picking_type_id)
        else:
            picking_type_id = self.env['stock.picking.type'].search([('code', '=', picking_type)], limit=1)
            print("----picking_type_id---_get_picking_type--------", picking_type_id)

        if not picking_type_id:
            raise UserError(_("Picking Type '%s' not found.") % picking_type)

        return picking_type_id.id
