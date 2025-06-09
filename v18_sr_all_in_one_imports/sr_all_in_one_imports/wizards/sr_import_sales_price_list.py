# -*- coding: utf-8 -*-
##############################################################################
#
#    OpenERP, Open Source Management Solution
#    Copyright (C) Sitaram Solutions (<https://sitaramsolutions.in/>).
#
#    For Module Support : info@sitaramsolutions.in  or Skype : contact.hiren1188
#
##############################################################################

import io
import csv
import base64
import openpyxl
from io import BytesIO
from datetime import datetime
from odoo.exceptions import UserError
from odoo import models, fields, api, _


class ImportPriceList(models.TransientModel):
    _name = 'import.price.list'
    _description = 'Import Sales Price List'

    file = fields.Binary('File', required=True)
    import_option = fields.Selection([('csv', 'CSV File'), ('xls', 'XLS File')], string='Select File Type',
                                     default='csv', required=True)
    import_type = fields.Selection([('product', 'Product'), ('category', 'Category')], string="Import Type",
                                   default='category', required=True)
    import_product_by = fields.Selection([('name', 'Name'), ('code', 'Code'), ('barcode', 'Barcode')],
                                         string='Import Product By', default='name')

    @api.onchange('import_type')
    def _onchange_import_type(self):
        if self.import_type == 'category':
            self.import_product_by = False
        else:
            self.import_product_by = 'name'

    def action_import_price_list(self):
        if not self.file:
            raise UserError(_('Please upload a file to proceed.'))
        if self.import_option == 'csv':
            if not self._is_csv_file():
                raise UserError(
                    _('You selected "CSV File" but uploaded a non-CSV file. Please upload a valid CSV file.'))
            self._import_price_list_csv()
        elif self.import_option == 'xls':
            if not self._is_xls_file():
                raise UserError(
                    _('You selected "XLS File" but uploaded a non-XLS file. Please upload a valid XLSX file.'))
            self._import_price_list_xls()
        else:
            raise UserError(_('Please select a valid file type to import.'))

    def _is_csv_file(self):
        try:
            data = base64.b64decode(self.file)
            file_content = data.decode('utf-8').splitlines()
            csv.reader(file_content, delimiter=',')
            return True
        except Exception:
            return False

    def _is_xls_file(self):
        try:
            data = base64.b64decode(self.file)
            BytesIO(data).seek(0)
            openpyxl.load_workbook(BytesIO(data))
            return True
        except Exception:
            return False

    def _import_price_list_csv(self):
        data = base64.b64decode(self.file)
        file_content = data.decode('utf-8').splitlines()
        reader = csv.reader(file_content, delimiter=',')
        next(reader)
        prev_pricelist_name = None
        prev_currency_name = None
        prev_country_group = None

        for row in reader:
            pricelist_name = row[0].strip() if row[0] else prev_pricelist_name
            currency_name = row[1].strip() if row[1] else prev_currency_name
            country_group = row[2].strip() if row[2] else prev_country_group
            product_name = row[3].strip() if row[3] else ''
            product_variant = row[4].strip() if row[4] else ''
            product_sku = row[5].strip() if row[5] else ''
            product_category = row[6].strip() if row[6] else ''
            price_type = row[7].strip() if row[7] else ''
            fixed_price = float(row[8]) if row[8] else 0.0
            percentage_price = float(row[9]) if row[9] else 0.0
            min_quantity = int(row[10]) if row[10] else 1
            date_start = row[11].strip() if row[11] else False
            date_end = row[12].strip() if row[12] else False

            if self.import_product_by == 'name' and not product_name:
                raise UserError(_('Product name is required when selecting "Name" for Import Product By.'))
            if self.import_product_by == 'code' and not product_variant:
                raise UserError(_('Product code is required when selecting "Code" for Import Product By.'))
            if self.import_product_by == 'barcode' and not product_sku:
                raise UserError(_('Product barcode is required when selecting "Barcode" for Import Product By.'))
            if self.import_type == 'category' and not product_category:
                raise UserError(_('Product category is required when importing by category.'))

            if row[0]:
                prev_pricelist_name = pricelist_name
                prev_currency_name = currency_name
                prev_country_group = country_group

            if not currency_name:
                raise UserError(_('Currency is required for the pricelist "%s".') % pricelist_name)

            currency = self.env['res.currency'].search([('name', '=', currency_name)], limit=1)
            if not currency:
                raise UserError(_('Currency "%s" not found.') % currency_name)

            if date_start:
                try:
                    date_start = datetime.strptime(date_start, '%m/%d/%y').date()
                except ValueError:
                    raise UserError(_('Invalid start date format for row "%s". Expected format: MM/DD/YY.') % row)
            if date_end:
                try:
                    date_end = datetime.strptime(date_end, '%m/%d/%y').date()
                except ValueError:
                    raise UserError(_('Invalid end date format for row "%s". Expected format: MM/DD/YY.') % row)

            self._process_row([
                pricelist_name, currency_name, country_group, product_name,
                product_variant, product_sku, product_category, price_type,
                fixed_price, percentage_price, min_quantity, date_start, date_end
            ])

    def _import_price_list_xls(self):
        data = base64.b64decode(self.file)
        workbook = openpyxl.load_workbook(BytesIO(data))
        sheet = workbook.active

        prev_pricelist_name = None
        prev_currency_name = None
        prev_country_group = None

        for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if all(cell is None for cell in row):
                continue

            pricelist_name = row[0] or prev_pricelist_name
            currency_name = row[1] or prev_currency_name
            country_group = row[2] or prev_country_group
            product_name = row[3] or ''
            product_variant = row[4] or ''
            product_sku = row[5] or ''
            product_category = row[6] or ''
            price_type = row[7] or ''
            fixed_price = row[8] or 0.0
            percentage_price = row[9] or 0.0
            min_quantity = row[10] or 1
            date_start = row[11] or False
            date_end = row[12] or False

            if self.import_product_by == 'name' and not product_name:
                raise UserError(_('Product name is required when selecting "Name" for Import Product By.'))
            if self.import_product_by == 'code' and not product_variant:
                raise UserError(_('Product code is required when selecting "Code" for Import Product By.'))
            if self.import_product_by == 'barcode' and not product_sku:
                raise UserError(_('Product barcode is required when selecting "Barcode" for Import Product By.'))
            if self.import_type == 'category' and not product_category:
                raise UserError(_('Product category is required when importing by category.'))

            if row[0]:
                prev_pricelist_name = pricelist_name
                prev_currency_name = currency_name
                prev_country_group = country_group

            if currency_name:
                currency = self.env['res.currency'].search([('name', '=', currency_name)], limit=1)
                if not currency:
                    raise UserError(_('Currency "%s" not found.') % currency_name)
            else:
                raise UserError(_('Currency is required for the pricelist "%s".') % pricelist_name)

            self._process_row([
                pricelist_name, currency_name, country_group, product_name,
                product_variant, product_sku, product_category, price_type,
                fixed_price, percentage_price, min_quantity, date_start, date_end
            ])

    def _process_row(self, row):
        pricelist_name = row[0] or ''
        currency_name = row[1] or ''
        country_group = row[2] or ''
        product_name = row[3] or ''
        product_variant = row[4] or ''
        product_sku = row[5] or ''
        product_category = row[6] or ''
        price_type = row[7] or ''
        fixed_price = row[8] or 0.0
        percentage_price = row[9] or 0.0
        min_quantity = row[10] or 1
        date_start = row[11] or False
        date_end = row[12] or False
        product_category_id = False

        pricelist = self.env['product.pricelist'].search([('name', '=', pricelist_name)], limit=1)
        if not pricelist:
            pricelist = self.env['product.pricelist'].create({
                'name': pricelist_name,
                'currency_id': self.env['res.currency'].search([('name', '=', currency_name)],
                                                               limit=1).id if currency_name else None,
            })

        if not pricelist.currency_id:
            raise UserError(_('Currency is missing for pricelist "%s".') % pricelist_name)

        if country_group:
            country_group_rec = self.env['res.country.group'].search([('name', '=', country_group)], limit=1)
            if not country_group_rec:
                raise UserError(_('Country Group "%s" not found.') % country_group)
            pricelist.country_group_ids = [(6, 0, [country_group_rec.id])]

        product = False
        if self.import_type == 'product':
            product = self._search_product(product_name, product_sku)

        elif self.import_type == 'category':
            product_category_rec = self.env['product.category'].search([('display_name', '=', product_category)],
                                                                       limit=1)
            if not product_category_rec:
                raise UserError(_('Product Category "%s" not found.') % product_category)
            product_category_id = product_category_rec.id
        else:
            raise UserError(_('Invalid Import Type.'))

        if price_type.lower() == 'fixed':
            compute_price = 'fixed'
            price_value = float(fixed_price)
        elif price_type.lower() == 'percentage':
            compute_price = 'percentage'
            price_value = float(percentage_price)
        else:
            raise UserError(_('Invalid Price Type "%s". Use "fixed" or "percentage".') % price_type)

        self.env['product.pricelist.item'].create({
            'pricelist_id': pricelist.id,
            'product_tmpl_id': product.product_tmpl_id.id if product else False,
            'product_id': product.id if product else False,
            'categ_id': product_category_id if self.import_type == 'category' else False,
            'compute_price': compute_price,
            'fixed_price': price_value if compute_price == 'fixed' else 0.0,
            'percent_price': price_value if compute_price == 'percentage' else 0.0,
            'min_quantity': min_quantity,
            'date_start': date_start,
            'date_end': date_end,
        })

    def _search_product(self, name, sku):
        domain = []
        if self.import_product_by == 'name':
            domain = [('name', '=', name)]
        elif self.import_product_by == 'code':
            domain = [('default_code', '=', sku)]
        elif self.import_product_by == 'barcode':
            domain = [('barcode', '=', sku)]

        product = self.env['product.product'].search(domain, limit=1)
        if not product:
            raise UserError(_('Product not found with %s: "%s"') % (self.import_product_by, sku or name))
        return product
