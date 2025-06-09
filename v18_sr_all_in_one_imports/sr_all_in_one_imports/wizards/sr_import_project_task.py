# -*- coding: utf-8 -*-
##############################################################################
#
#    OpenERP, Open Source Management Solution
#    Copyright (C) Sitaram Solutions (<https://sitaramsolutions.in/>).
#
#    For Module Support : info@sitaramsolutions.in  or Skype : contact.hiren1188
#
##############################################################################

import csv
import base64
import tempfile
from datetime import datetime
from odoo import models, fields, _
from openpyxl import load_workbook
from odoo.exceptions import UserError


class srImportProjectTask(models.TransientModel):
    _name = 'import.project.task'
    _description = 'Import Project Task'

    file = fields.Binary('File')
    import_option = fields.Selection([('csv', 'CSV File'), ('xls', 'XLS File')], string='Select', default='csv')

    def import_project_task(self):
        if not self.file:
            raise UserError(_("Please upload a file to import."))

        data = base64.b64decode(self.file)
        if self.import_option == 'csv':
            self._import_csv(self.file)
        elif self.import_option == 'xls':
            self._import_xls(data)

    def _import_csv(self, data):
        try:
            if isinstance(data, bytes):
                file_content = base64.b64decode(data).decode("utf-8")
            else:
                file_content = data.decode("utf-8")

            csv_reader = csv.reader(file_content.splitlines(), delimiter=',')
            headers = next(csv_reader)

            for row in csv_reader:
                row_data = dict(zip(headers, row))
                self._create_project_task(row_data)
        except Exception as e:
            raise UserError(_("Error processing CSV file: %s") % str(e))

    def _import_xls(self, data):
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                tmp.write(data)
                tmp.flush()
                workbook = load_workbook(tmp.name, data_only=True)

            sheet = workbook.active
            headers = [cell.value for cell in sheet[1]]

            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_data = dict(zip(headers, row))
                self._create_project_task(row_data)
        except Exception as e:
            raise UserError(_("Error processing XLSX file: %s") % str(e))

    def _create_project_task(self, values):
        task_name = values.get('Task Name')

        project_name = values.get('Project')
        project_id = False
        if project_name:
            project_id = self.env['project.project'].search([('name', '=', project_name)], limit=1).id
            if not project_id and project_name:
                raise UserError(_("Project '%s' not found") % project_name)

        assignee_ids = []
        if values.get('Assignees') and values['Assignees']:
            assignee_names = str(values.get('Assignees', '')).split(',')
            for assignee_name in assignee_names:
                user = self.env['res.users'].search([('name', '=', assignee_name.strip())], limit=1)
                if not user:
                    raise UserError(_("Assignee user '%s' not found") % assignee_name)
                assignee_ids.append(user.id)

        tag_ids = []
        if values.get('Tags') and values['Tags']:
            tag_names = str(values.get('Tags', '')).split(',')
            for tag_name in tag_names:
                tag = self.env['project.tags'].search([('name', '=', tag_name.strip())], limit=1)
                if not tag:
                    raise UserError(_("Tags '%s' not found") % tag_name)
                tag_ids.append(tag.id)
        try:
            deadline = self._parse_date(values.get('Deadline'))
        except ValueError as e:
            raise UserError(_("Invalid date format: %s") % str(e))

        parent_task = values.get('Parent Task')
        parent_task_id = False
        if parent_task:
            parent_task_id = self.env['project.task'].search([('name', '=', parent_task)], limit=1).id
            if not parent_task_id and parent_task:
                raise UserError(_("Parent task '%s' not found") % parent_task)

        if task_name:
            task_vals = {
                'name': task_name,
                'project_id':  project_id if project_id else False,
                'user_ids': [(6, 0, assignee_ids)],
                'tag_ids': [(6, 0, tag_ids)],
                'date_deadline' : deadline,
                'parent_id': parent_task_id if parent_task_id else False,
                'sequence': values.get('Sequence', ''),
                'description': values.get('Description', ''),
            }
            task_id = self.env['project.task'].create(task_vals)
        else:
            raise UserError(_("Name must be required.!"))

    def _parse_date(self, date_value):
        if not date_value:
            return None

        if isinstance(date_value, datetime):
            return date_value.date()
        try:
            return datetime.strptime(date_value, '%Y-%m-%d').date()
        except ValueError:
            raise ValueError(f"Date '{date_value}' is in an invalid format. Expected YYYY-MM-DD.")
