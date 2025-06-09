# -*- coding: utf-8 -*-
##############################################################################
#
#    OpenERP, Open Source Management Solution
#    Copyright (C) Sitaram Solutions (<https://sitaramsolutions.in/>).
#
#    For Module Support : info@sitaramsolutions.in  or Skype : contact.hiren1188
#
##############################################################################

import io
import csv
import xlrd
import base64
import logging
import tempfile
import binascii
from datetime import datetime
from odoo import models, fields, _
from openpyxl import load_workbook
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

class srImportStockPicking(models.TransientModel):
    _name = 'import.stock.picking'
    _description = 'Import Stock Picking'

    file = fields.Binary('File')
    import_option = fields.Selection([('csv', 'CSV File'), ('xls', 'XLS File')], string='Select', default='csv')
    picking_stage_selection = fields.Selection([('draft', 'Draft'), ('confirm', 'Confirm picking automatically when import')], string='Picking Stage', default='draft')
    sequence_option = fields.Selection([('custom', 'Use sequence from Excel/CSV'), ('default', 'Use Default Sequence')], string='Sequnce option', default='default')
    import_product_by = fields.Selection([('name', 'Name'), ('default_code', 'Code'), ('barcode', 'Barcode')], string='Import Product By', default='name')
    picking_type = fields.Selection([('incoming', 'Receipts'), ('outgoing', 'Deliveries'), ('internal', 'Internal')], string='Picking Type', default='incoming')
    source_location_id = fields.Many2one('stock.location', string="Source Location")
    destination_location_id = fields.Many2one('stock.location', string="Destination Location")

    def import_stock_picking(self):
        if not self.file:
            raise UserError(_("Please upload a file to import."))

        data = base64.b64decode(self.file)
        if self.import_option == 'csv':
            self._import_csv(self.file)
        elif self.import_option == 'xls':
            self._import_xls(data)

    def _import_csv(self, data):
        try:
            if isinstance(data, bytes):
                file_content = base64.b64decode(data).decode("utf-8")
            else:
                file_content = data.decode("utf-8")

            csv_reader = csv.reader(file_content.splitlines(), delimiter=',')
            headers = next(csv_reader)
            _logger.info("Detected Headers: %s", headers)

            last_name = None
            pickings = {}
            Picking = self.env['stock.picking']

            for row in csv_reader:
                row_data = dict(zip(headers, row))
                _logger.info("Row Data: %s", row_data)

                name = row_data.get('Name') or last_name
                source_document = row_data.get('Source Document')
                date = row_data.get('Date') or row_data.get('date')
                partner = row_data.get('Contact')

                if not name:
                    raise UserError(_("Missing 'Name' value in file. Unable to proceed."))

                last_name = name
                if name not in pickings:
                    pickings[name] = self._create_picking(name, source_document, date, partner)

                self._create_move(pickings[name], row_data)

            if self.picking_stage_selection == 'confirm':
                for name, picking in pickings.items():
                    picking_name = picking.name
                    picking_obj = Picking.search([('name', '=', picking_name)], limit=1)
                    if picking_obj:
                        picking_obj.button_validate()
        except Exception as e:
            raise UserError(_("Error processing CSV file: %s") % str(e))

    def _import_xls(self, data):
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                tmp.write(data)
                tmp.flush()
                workbook = load_workbook(tmp.name, data_only=True)

            sheet = workbook.active
            headers = [cell.value for cell in sheet[1]]  # Read first row as headers

            _logger.info("Detected Headers: %s", headers)

            pickings = {}
            last_name = None
            Picking = self.env['stock.picking']

            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_data = dict(zip(headers, row))
                _logger.info("Row Data: %s", row_data)

                name = row_data.get('Name') or last_name
                source_document = row_data.get('Source Document')
                date = row_data.get('Date') or row_data.get('date')
                partner = row_data.get('Contact')

                if not name:
                    raise UserError(_("Missing 'Name' value in file. Unable to proceed."))

                last_name = name
                if name not in pickings:
                    pickings[name] = self._create_picking(name, source_document, date, partner)
                self._create_move(pickings[name], row_data)

            if self.picking_stage_selection == 'confirm':
                for name, picking in pickings.items():
                    picking_name = picking.name
                    picking_obj = Picking.search([('name', '=', picking_name)], limit=1)
                    if picking_obj:
                        picking_obj.button_validate()

        except Exception as e:
            raise UserError(_("Error processing XLSX file: %s") % str(e))

    def _create_picking(self, name, source_document, date, partner):
        Picking = self.env['stock.picking']

        picking_vals = {
            'picking_type_id': self._get_picking_type(),
            'partner_id' : self._find_partner(partner),
            'origin': source_document or f"Import Order {name}",
            'state': 'draft' if self.picking_stage_selection == 'draft' else 'confirmed',
        }
        if self.sequence_option == 'custom':
            if self.picking_type == 'incoming':
                picking_vals.update({
                    'name': name,
                    'location_dest_id': self.destination_location_id.id,
                })
            elif self.picking_type == 'outgoing':
                picking_vals.update({
                    'name': name,
                    'location_id': self.source_location_id.id,
                })
            else:
                picking_vals.update({
                    'name': name,
                    'location_id': self.source_location_id.id,
                    'location_dest_id': self.destination_location_id.id,
                })
        else:
            if self.picking_type == 'incoming':
                picking_vals.update({
                    'location_dest_id': self.destination_location_id.id,
                })
            elif self.picking_type == 'outgoing':
                picking_vals.update({
                    'location_id': self.source_location_id.id,
                })
            else:
                picking_vals.update({
                    'location_id': self.source_location_id.id,
                    'location_dest_id': self.destination_location_id.id,
                })
        picking = Picking.create(picking_vals)
        return picking

    def _find_partner(self, customer):
        if not customer:
            return False
        partner_id = self.env['res.partner'].search([('name', '=', customer)])
        if partner_id:
            return partner_id.id

    def _create_move(self, picking, row):
        Move = self.env['stock.move']
        product = self._find_product(row)
        stock_lot_obj = self.env['stock.lot']

        if row.get('Date') and row['Date']:
            scheduled_date = datetime.strptime(row.get('Date'), '%d/%m/%Y')
        else:
            scheduled_date = datetime.now()

        move_vals = {
            'picking_id': picking.id,
            'product_id': product.id,
            'name': product.name,
            'date': scheduled_date,
            'product_uom_qty': float(row.get('Quantity', 1)),
            'product_uom': product.uom_id.id,
            'location_id': picking.location_id.id,
            'location_dest_id': picking.location_dest_id.id,
            'price_unit': float(row.get('Price', 0)),
        }
        move_id = Move.create(move_vals)
        picking.action_confirm()
        lot_name = row.get('Lot')
        if lot_name:
            lot_id = stock_lot_obj.sudo().create({
                'name': lot_name,
                'product_id': product.id,
            })
            if lot_id and move_id:
                for move_line in move_id.move_line_ids:
                    move_line.sudo().write({
                        'lot_id': lot_id.id,
                    })

    def _find_product(self, row):
        Product = self.env['product.product']
        search_field = self.import_product_by
        row = {key.lower(): value for key, value in row.items()}
        search_value = row.get('product name') or row.get('Product Name')
        if not search_value:
            raise UserError(_("Missing product %s value in file. Available columns: %s") % (search_field, row.keys()))
        product = Product.search([(search_field, '=', search_value)], limit=1)
        if not product:
            raise UserError(_("Product with %s '%s' not found. Check product records.") % (search_field, search_value))
        return product

    def _get_picking_type(self):
        picking_type = self.picking_type
        if picking_type == 'internal':
            picking_type_id = self.env['stock.picking.type'].search([('code', '=', picking_type), ('sequence_code', '=', 'INT')], limit=1)
        else:
            picking_type_id = self.env['stock.picking.type'].search([('code', '=', picking_type)], limit=1)
        if not picking_type_id:
            raise UserError(_("Picking Type '%s' not found.") % picking_type)
        return picking_type_id.id
