# -*- coding: utf-8 -*-
##############################################################################
#
#    OpenERP, Open Source Management Solution
#    Copyright (C) Sitaram Solutions (<https://sitaramsolutions.in/>).
#
#    For Module Support : info@sitaramsolutions.in  or Skype : contact.hiren1188
#
##############################################################################

import io
import csv
import xlrd
import base64
import binascii
import tempfile
from odoo import models, fields, _
from openpyxl import load_workbook
from odoo.exceptions import UserError


class srImportpurchaseOrderLines(models.TransientModel):
    _name = 'sr.import.purchase.order.lines'
    _description = 'Import Purchase Order Lines'

    file = fields.Binary('File')
    import_option = fields.Selection([('csv', 'CSV File'), ('xls', 'XLS File')], string='Select File Type', default='csv')
    use_product_details_from = fields.Selection([('custom', 'Use Product Details from Excel/CSV'), ('from_product', 'Use Product Details from Product Configurations')], string='Use Product Details Option', default='custom')
    import_product_by = fields.Selection([('name', 'Name'), ('code', 'code'), ('barcode', 'barcode')], string='Import Product By', default='name')

    def _find_product(self, product):
        if self.import_product_by == 'name':
            product_id = self.env['product.product'].search([('name', '=', product)])
            if not product_id:
                raise UserError(_('Product with name %s does not exist in system' % product))
        elif self.import_product_by == 'code':
            product_id = self.env['product.product'].search([('default_code', '=', product)])
            if not product_id:
                raise UserError(_('Product with code %s does not exist in system' % product))
        else:
            product_id = self.env['product.product'].search([('barcode', '=', product)])
            if not product_id:
                raise UserError(_('Product with barcode %s does not exist in system' % product))
        return product_id
    
    def _find_uom(self, uom):
        uom_id = self.env['uom.uom'].search([('name', '=', uom)])
        if uom_id:
            return uom_id.id
        else:
            raise UserError(_('%s Unit Of Measure does not exist in system' % uom))

    def generate_purchase_order_line(self, line):
        product_ids = self._find_product(line[0])
        if line[5]:
            tax_ids = []
            for record in line[5].split(','):
                tax_id = self.env['account.tax'].search([('name', '=', record), ('type_tax_use', '=', 'purchase')])
                if not tax_id:
                    raise Warning(_('"%s" Tax not available in your system') % record)
                tax_ids.append(tax_id.id)
                
        if product_ids:
            for product_id in product_ids:
                self.env['purchase.order.line'].create({
                    'product_id':product_id.id,
                    'name':product_id.description_purchase or product_id.name if self.use_product_details_from == 'from_product' else line[1],
                    'product_qty':line[2],
                    'product_uom':product_id.uom_id.id if self.use_product_details_from == 'from_product' else self._find_uom(line[3]),
                    'price_unit':product_id.list_price if self.use_product_details_from == 'from_product' else line[4],
                    'order_id':self._context.get('active_id'),
                    'taxes_id': ([(6, 0, product_id.supplier_taxes_id.ids)]) if self.use_product_details_from == 'from_product' else ([(6, 0, tax_ids)])
                    })

    def import_purchase_order_lines(self):
        try:
            if self.import_option == 'xls':
                temp_excel_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
                temp_excel_file.write(binascii.a2b_base64(self.file))
                temp_excel_file.seek(0)
                workbook = load_workbook(temp_excel_file.name, data_only=True)
                sheet = workbook.active
                
                for row_no, row in enumerate(sheet.iter_rows(values_only=True)):
                    if row_no == 0:
                        continue
                    line = [cell for cell in row]
                    self.generate_purchase_order_line(line)
                
                temp_excel_file.close()
            else:
                csv_data = base64.b64decode(self.file)
                data_file = io.StringIO(csv_data.decode("utf-8"))
                csv_reader = csv.reader(data_file, delimiter=',')
                
                for row_no, line in enumerate(csv_reader):
                    if row_no == 0:
                        continue
                    self.generate_purchase_order_line(line)
        except Exception as e:
            raise UserError(_("An error occurred: %s") % str(e))