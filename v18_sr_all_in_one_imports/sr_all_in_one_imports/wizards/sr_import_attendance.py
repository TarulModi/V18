# -*- coding: utf-8 -*-
##############################################################################
#
#    OpenERP, Open Source Management Solution
#    Copyright (C) Sitaram Solutions (<https://sitaramsolutions.in/>).
#
#    For Module Support : info@sitaramsolutions.in  or Skype : contact.hiren1188
#
##############################################################################

import io
import csv
import xlrd
import base64
import tempfile
import binascii
from datetime import datetime
from openpyxl import load_workbook
from odoo.exceptions import UserError
from odoo import models, fields, api, _

import logging

_logger = logging.getLogger(__name__)

class ImportEmployeeAttendance(models.TransientModel):
    _name = 'sr.hr.import.attendance'
    _description = 'Import Employee Attendance'

    import_file = fields.Binary('Select File')
    import_file_by = fields.Selection([('csv', 'CSV File'), ('xls', 'XLS File')], string='Select File Type', default='csv')
    import_employee_option = fields.Selection([('name', 'Name'), ('badge', 'Badge')], string='Import Employee By', default='name')

    def _import_attendance(self, line):
        if not line[0]:
            raise Warning('Please provide employee name or badge')
        if not line[1]:
            raise Warning('Please provide sign_in Date and time')

        if self.import_employee_option == 'name':
            employee_id = self.env['hr.employee'].search([('name', '=', line[0])])
        else:
            try:
                employee_id = self.env['hr.employee'].search([('barcode', '=', int(float(line[0])))])
            except ValueError:
                raise Warning(f"Invalid badge number: {line[0]}")

        if not employee_id:
            raise Warning('Employee Not Found in System')

        if not line[2]:
            self.env['hr.attendance'].create({
                'employee_id': employee_id.id, 'check_in': line[1]
            })
        else:
            self.env['hr.attendance'].create({
                'employee_id': employee_id.id, 'check_in': line[1], 'check_out': line[2]
            })

        return True


    def import_employee_attendance(self):
        try:
            if self.import_file_by == 'xls':
                try:
                    temp_excel_file = tempfile.NamedTemporaryFile(suffix=".xlsx")
                    _logger.info("Attempting to write base64 file data to temp file.")
                    temp_excel_file.write(binascii.a2b_base64(self.import_file))
                    temp_excel_file.seek(0)
                    workbook = load_workbook(temp_excel_file.name, data_only=True)
                    sheet = workbook.active

                    for row_no, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                        if row_no != 1:
                            line = list(map(lambda cell: str(cell) if cell is not None else '', row))
                            if line[1]:
                                line[1] = datetime.strptime(str(line[1]), '%Y-%m-%d %H:%M:%S') if isinstance(line[1], str) else line[1]
                            if line[2]:
                                line[2] = datetime.strptime(str(line[2]), '%Y-%m-%d %H:%M:%S') if isinstance(line[2], str) else line[2]
                            self._import_attendance(line)
                except Exception as e:
                    _logger.error(f"Error reading Excel file: {str(e)}")
                    raise UserError(str(e))
            else:
                try:
                    csv_data = base64.b64decode(self.import_file)
                    data_file = io.StringIO(csv_data.decode("utf-8"))
                    data_file.seek(0)
                    file_reader = []
                    csv_reader = csv.reader(data_file, delimiter=',')
                    file_reader.extend(csv_reader)
                    _logger.info(f"Processing {len(file_reader)} rows in the CSV file")
                except Exception as e:
                    _logger.error(f"Error reading CSV file: {str(e)}")
                    raise UserError("Invalid file format or error while reading the file")
                for row_no in range(len(file_reader)):
                    line = list(map(str, file_reader[row_no]))
                    if line:
                        if row_no != 0:
                            self._import_attendance(line)
        except Exception as e:
            _logger.error(f"General error during import: {str(e)}")
            raise UserError(str(e))
