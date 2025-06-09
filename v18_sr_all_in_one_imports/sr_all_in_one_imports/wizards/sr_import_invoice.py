# -*- coding: utf-8 -*-
##############################################################################
#
#    OpenERP, Open Source Management Solution
#    Copyright (C) Sitaram Solutions (<https://sitaramsolutions.in/>).
#
#    For Module Support : info@sitaramsolutions.in  or Skype : contact.hiren1188
#
##############################################################################

from odoo import models, fields, api, _
from odoo.exceptions import UserError
from datetime import datetime
import tempfile
import binascii
import xlrd
import base64
import io
import csv
from openpyxl import load_workbook


TYPE2JOURNAL = {
    'customer': 'sale',
    'vendor': 'purchase',
    'customer_refund': 'sale',
    'vendor_refund': 'purchase',
}


class ImportInvoice(models.TransientModel):
    _name = 'import.invoice'
    _description = 'Import Invoice'

    file = fields.Binary('File')
    import_option = fields.Selection([('csv', 'CSV File'), ('xls', 'XLS File')], string='Select File', default='csv')
    invoice_stage_selection = fields.Selection([('draft', 'Draft Invoice'), ('open', 'Validate invoice')], string='Invoice Stage', default='draft')
    sequence_option = fields.Selection([('custom', 'Use sequence from Excel/CSV'), ('default', 'Use Default Sequence')], string='Sequnce option', default='default')
    import_product_by = fields.Selection([('name', 'Name'), ('code', 'Code'), ('barcode', 'Barcode')], string='Import Product By', default='name')
    import_customer_by = fields.Selection([('name', 'Name'), ('ref', 'Internal Reference')], string='Import Customer/Vendor By', default='name')
    import_account_by = fields.Selection([('file', 'From Excel/CSV File'), ('product', 'From Product')], string='Import Invoice Line Account By', default='product')
    import_invoice_by = fields.Selection([('customer', 'Customer Invoice'), ('vendor', 'Vendor Invoice'), ('customer_refund', 'Customer Refund'), ('vendor_refund', 'Vendor Refund')], string='Import Invoice By', default='customer')
    
    def _find_partner(self, customer):
        if not customer:
            return False
        if self.import_customer_by == 'name':
            partner_id = self.env['res.partner'].search([('name', '=', customer)])
        else:
            partner_id = self.env['res.partner'].search([('ref', '=', customer)])
        if partner_id:
            return partner_id.id
        else:
            raise UserError(_('%s Customer does not exist in system' % customer))

    def _find_user(self, salesperson):
        if not salesperson:
            return False
        user_id = self.env['res.users'].search([('name', '=', salesperson)])
        if user_id:
            return user_id.id
        else:
            raise UserError(_('%s salesperson does not exist in system' % salesperson))

    def _find_product(self, product):
        if self.import_product_by == 'name':
            product_id = self.env['product.product'].search([('name', '=', product)])
            if not product_id:
                raise UserError(_('Product with name %s does not exist in system' % product))
        elif self.import_product_by == 'code':
            product_id = self.env['product.product'].search([('default_code', '=', product)])
            if not product_id:
                raise UserError(_('Product with code %s does not exist in system' % product))
        else:
            product_id = self.env['product.product'].search([('barcode', '=', product)])
            if not product_id:
                raise UserError(_('Product with barcode %s does not exist in system' % product))
        return product_id.id

    def _find_uom(self, uom):
        uom_id = self.env['uom.uom'].search([('name', '=', uom)])
        if uom_id:
            return uom_id.id
        else:
            raise UserError(_('%s Unit Of Measure does not exist in system' % uom))

    def _find_payment_term(self, payment_term):
        payment_term_id = self.env['account.payment.term'].search([('name', '=', payment_term)])
        if payment_term_id:
            return payment_term_id.id
        else:
            raise UserError(_('%s Payment Terms does not exist in system' % payment_term))

    def _find_account_id(self, account):
        account_id = self.env['account.account'].search([('name', '=', account)])
        if not account_id:
            raise UserError(_('%s Account does not exist in system' % account))
        return account_id.id
    
    def _find_product_account_id(self, product):
        if self.import_product_by == 'name':
            product_id = self.env['product.product'].search([('name', '=', product)])
            if not product_id:
                raise UserError(_('Product with name %s does not exist in system' % product))
        elif self.import_product_by == 'code':
            product_id = self.env['product.product'].search([('default_code', '=', product)])
            if not product_id:
                raise UserError(_('Product with code %s does not exist in system' % product))
        else:
            product_id = self.env['product.product'].search([('barcode', '=', product)])
            if not product_id:
                raise UserError(_('Product with barcode %s does not exist in system' % product))
        if self.import_invoice_by == 'customer' or self.import_invoice_by == 'customer_refund':
            return product_id.categ_id.property_account_income_categ_id.id
        elif self.import_invoice_by == 'vendor' or self.import_invoice_by == 'vendor_refund':
            return product_id.categ_id.property_account_expense_categ_id.id

    def generate_invoice_line(self, product, description, quantity, uom, price, account, tax):
        invoice_line_dict = {}
        tax_ids = []
        if tax:
            type_tax_use = ''
            for record in tax.split(','):
                if self.import_invoice_by == 'customer' or self.import_invoice_by == 'customer_refund':
                    type_tax_use = 'sale'
                else:
                    type_tax_use = 'purchase'
                tax_id = self.env['account.tax'].search([('name', '=', record), ('type_tax_use', '=', type_tax_use)])
                if not tax_id:
                    raise Warning(_('"%s" Tax not available in your system') % record)
                tax_ids.append(tax_id.id)
        
        if self.import_account_by == 'file':
            invoice_line_dict = {
                'product_id':self._find_product(product),
                'name':description,
                'quantity':quantity,
                'product_uom_id':self._find_uom(uom),
                'price_unit':price,
                'account_id': self._find_account_id(account),
                'tax_ids':([(6, 0, tax_ids)])
                }
        else:     
            invoice_line_dict = {
                'product_id':self._find_product(product),
                'name':description,
                'quantity':quantity,
                'product_uom_id':self._find_uom(uom),
                'price_unit':price,
                'account_id':self._find_product_account_id(product),
                'tax_ids':([(6, 0, tax_ids)])
                }
        return invoice_line_dict

    def _default_journal(self):
        inv_type = self.import_invoice_by
        inv_types = inv_type if isinstance(inv_type, list) else [inv_type]
        company_id = self._context.get('company_id', self.env.user.company_id.id)
        domain = [
            ('type', 'in', [TYPE2JOURNAL[ty] for ty in inv_types if ty in TYPE2JOURNAL]),
            ('company_id', '=', company_id),
        ]
        return self.env['account.journal'].search(domain, limit=1)

    def generate_invoice(self, line):
        invoice_ids = []
        invoice_id = self.env['account.move'].search([('name', '=', line[0])])
        if invoice_id:
            if invoice_id.partner_id.name == line[1]:
                if invoice_id.user_id.name == line[4]:
                    invoice_id.write({'invoice_line_ids':[(0, 0, self.generate_invoice_line(product=line[5], description=line[6], quantity=line[8], uom=line[9], price=line[10], account=line[7], tax=line[11]))]})
                else:
                    raise UserError(_('Sales Person is different for %s, \n Please define same' % invoice_id.name))
            else:
                raise UserError(_('Customer is different for %s, \n Please define same' % invoice_id.name))
        else:
            if self.import_invoice_by == 'customer':
                type = 'out_invoice'
            elif self.import_invoice_by == 'vendor':
                type = 'in_invoice'
            elif self.import_invoice_by == 'customer_refund':
                type = 'out_refund'
            else:
                type = 'in_refund'
            
            invoice_id = self.env['account.move'].create({
                'name':line[0] if self.sequence_option == 'custom' else '/',
                'partner_id':self._find_partner(line[1]),
                'invoice_date':line[2],
                'user_id':self._find_user(line[4]),
                'invoice_payment_term_id':self._find_payment_term(line[3]),
                'move_type':type,
                'journal_id': self._default_journal().id,
                'invoice_line_ids':[(0, 0, self.generate_invoice_line(product=line[5], description=line[6], quantity=line[8], uom=line[9], price=line[10], account=line[7], tax=line[11]))]
                })
            invoice_ids.append(invoice_id.id)
        if self.invoice_stage_selection == 'open':
            self.env['account.move'].browse(invoice_ids).action_post()
        
    def import_invoices(self):
        try:
            if self.import_option == 'xls':
                temp_excel_file = tempfile.NamedTemporaryFile(suffix=".xlsx")
                temp_excel_file.write(binascii.a2b_base64(self.file))
                temp_excel_file.seek(0)
                workbook = load_workbook(temp_excel_file.name)
                sheet = workbook.active

                for row_no, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                    if row_no != 1:
                        line = [str(cell) if cell is not None else '' for cell in row]
                        self.generate_invoice(line)
            else:
                try:
                    csv_data = base64.b64decode(self.file)
                    data_file = io.StringIO(csv_data.decode("utf-8"))
                    data_file.seek(0)
                    file_reader = []
                    csv_reader = csv.reader(data_file, delimiter=',')
                    file_reader.extend(csv_reader)
                except Exception as decode_error:
                    raise UserError(_("Invalid file! Error: %s") % str(decode_error))
                for row_no in range(len(file_reader)):
                    line = list(map(str, file_reader[row_no]))
                    if line:
                        if row_no != 0:
                            self.generate_invoice(line)
        except Exception as e:
            raise UserError(_('An error occurred: %s') % str(e))
