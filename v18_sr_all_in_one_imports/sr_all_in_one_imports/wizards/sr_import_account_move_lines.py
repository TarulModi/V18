# -*- coding: utf-8 -*-
##############################################################################
#
#    OpenERP, Open Source Management Solution
#    Copyright (C) Sitaram Solutions (<https://sitaramsolutions.in/>).
#
#    For Module Support : info@sitaramsolutions.in  or Skype : contact.hiren1188
#
##############################################################################

import io
import csv
import xlrd
import base64
import binascii
import tempfile
from odoo import models, fields, _
from openpyxl import load_workbook
from odoo.exceptions import UserError



class srImportmoveLines(models.TransientModel):
    _name = 'sr.import.account.move.lines'
    _description = 'Import Account Move Lines'

    file = fields.Binary('File')
    import_option = fields.Selection([('csv', 'CSV File'), ('xls', 'XLS File')], string='Select File Type', default='csv')
    use_product_details_from = fields.Selection([('custom', 'Use Product Details from Excel/CSV'), ('from_product', 'Use Product Details from Product Configurations')], string='Use Product Details Option', default='custom')
    import_product_by = fields.Selection([('name', 'Name'), ('code', 'code'), ('barcode', 'barcode')], string='Import Product By', default='name')

    def _find_product(self, product):
        if self.import_product_by == 'name':
            product_id = self.env['product.product'].search([('name', '=', product)])
            if not product_id:
                raise UserError(_('Product with name %s does not exist in system' % product))
        elif self.import_product_by == 'code':
            product_id = self.env['product.product'].search([('default_code', '=', product)])
            if not product_id:
                raise UserError(_('Product with code %s does not exist in system' % product))
        else:
            product_id = self.env['product.product'].search([('barcode', '=', product)])
            if not product_id:
                raise UserError(_('Product with barcode %s does not exist in system' % product))
        return product_id
    
    def _find_account_id(self, account):
        account_id = self.env['account.account'].search([('name', '=', account)])
        if account_id:
            return account_id.id
        else:
            raise UserError(_('%s Account does not exist in system' % account))

    def _find_uom(self, uom):
        uom_id = self.env['uom.uom'].search([('name', '=', uom)])
        if uom_id:
            return uom_id.id
        else:
            raise UserError(_('%s Unit Of Measure does not exist in system' % uom))

    def generate_account_move_line(self, line):
        product_ids = self._find_product(line[0])
        move_id = self.env['account.move'].browse(self._context.get('active_id'))
        if line[6]:
            tax_ids = []
            for record in line[6].split(','):
                if move_id.move_type == 'out_invoice':
                    tax_id = self.env['account.tax'].search([('name', '=', record), ('type_tax_use', '=', 'sale')])
                else:
                    tax_id = self.env['account.tax'].search([('name', '=', record), ('type_tax_use', '=', 'purchase')])
                if not tax_id:
                    raise Warning(_('"%s" Tax not available in your system') % record)
                tax_ids.append(tax_id.id)
        
        fiscal_position = move_id.fiscal_position_id
        if product_ids:
            for product_id in product_ids:
                accounts = product_id.product_tmpl_id.get_product_accounts(fiscal_pos=fiscal_position)
                if self.use_product_details_from == 'from_product' and move_id.move_type == 'out_invoice':
                    name = product_id.description_sale or product_id.name
                    unit_price = product_id.list_price
                    product_tax_ids = [(6, 0, product_id.taxes_id.ids)]
                    account_id = accounts['income']
                else:
                    name = product_id.description_purchase
                    unit_price = product_id.standard_price
                    product_tax_ids = [(6, 0, product_id.supplier_taxes_id.ids)]
                    account_id = accounts['expense']

                move_id.write({
                    'invoice_line_ids': [
                        (0, 0, {
                    'product_id':product_id.id,
                    'name':name if self.use_product_details_from == 'from_product'  else line[1],
                    'quantity':line[3],
                    'product_uom_id':product_id.uom_po_id.id if self.use_product_details_from == 'from_product' else self._find_uom(line[4]),
                    'price_unit':unit_price if self.use_product_details_from == 'from_product' else line[5],
                    'move_id':self._context.get('active_id'),
                    'tax_ids': (product_tax_ids) if self.use_product_details_from == 'from_product' else ([(6, 0, tax_ids)]),
                    'account_id':account_id.id if self.use_product_details_from == 'from_product' else self._find_account_id(line[2])
                    }),
                    ]
                })
        
    def import_account_move_lines(self):
        try:
            if self.import_option == 'xls':
                temp_excel_file = tempfile.NamedTemporaryFile(suffix=".xlsx")
                temp_excel_file.write(binascii.a2b_base64(self.file))
                temp_excel_file.seek(0)
                
                workbook = load_workbook(temp_excel_file.name, data_only=True)
                sheet = workbook.active
                
                for row_no, row in enumerate(sheet.iter_rows(values_only=True)):
                    if row_no == 0:
                        continue
                    # line = list(row)
                    line = [cell for cell in row]
                    self.generate_account_move_line(line)
            else:
                try:
                    csv_data = base64.b64decode(self.file)
                    data_file = io.StringIO(csv_data.decode("utf-8"))
                    data_file.seek(0)
                    file_reader = []
                    csv_reader = csv.reader(data_file, delimiter=',')
                    file_reader.extend(csv_reader)
                except Exception:
                    raise UserError(_("Invalid file!"))
                
                for row_no in range(len(file_reader)):
                    line = list(map(str, file_reader[row_no]))
                    if line:
                        if row_no != 0:
                            self.generate_account_move_line(line)
        except Exception as e:
            raise UserError(_("An error occurred: %s") % str(e))
