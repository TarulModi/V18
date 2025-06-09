# -*- coding: utf-8 -*-
##############################################################################
#
#    OpenERP, Open Source Management Solution
#    Copyright (C) Sitaram Solutions (<https://sitaramsolutions.in/>).
#
#    For Module Support : info@sitaramsolutions.in  or Skype : contact.hiren1188
#
##############################################################################

import io
import csv
import xlrd
import base64
import tempfile
import binascii
from datetime import datetime
from odoo import models, fields, _
from openpyxl import load_workbook
from odoo.exceptions import UserError



class srImportSalesOrder(models.TransientModel):
    _name = 'import.sales.order'
    _description = 'Import Sales Order'

    file = fields.Binary('File')
    import_option = fields.Selection([('csv', 'CSV File'), ('xls', 'XLS File')], string='Select', default='csv')
    quotation_stage_selection = fields.Selection([('draft', 'Draft Quotaion'), ('confirm', 'Confirm quotation automatically when import')], string='Quotation Stage', default='draft')
    sequence_option = fields.Selection([('custom', 'Use sequence from Excel/CSV'), ('default', 'Use Default Sequence')], string='Sequnce option', default='default')
    import_product_by = fields.Selection([('name', 'Name'), ('code', 'code'), ('barcode', 'barcode')], string='Import Product By', default='name')

    def _find_partner(self, customer):
        if not customer:
            return False
        partner_id = self.env['res.partner'].search([('name', '=', customer)])
        if partner_id:
            return partner_id.id
        else:
            raise UserError(_('%s Customer does not exist in system' % customer))

    def _find_pricelist(self, pricelist):
        if not pricelist:
            return False
        pricelist_id = self.env['product.pricelist'].search([('name', '=', pricelist)])
        if pricelist_id:
            return pricelist_id.id
        else:
            raise UserError(_('%s does not exist in system' % pricelist))

    def _find_user(self, salesperson):
        if not salesperson:
            return False
        user_id = self.env['res.users'].search([('name', '=', salesperson)])
        if user_id:
            return user_id.id
        else:
            raise UserError(_('%s salesperson does not exist in system' % salesperson))

    def _find_product(self, product):
        if self.import_product_by == 'name':
            product_id = self.env['product.product'].search([('name', '=', product)])
            if not product_id:
                raise UserError(_('Product with name %s does not exist in system' % product))
        elif self.import_product_by == 'code':
            product_id = self.env['product.product'].search([('default_code', '=', product)])
            if not product_id:
                raise UserError(_('Product with code %s does not exist in system' % product))
        else:
            product_id = self.env['product.product'].search([('barcode', '=', product)])
            if not product_id:
                raise UserError(_('Product with barcode %s does not exist in system' % product))
        return product_id.id
    
    def _find_uom(self, uom):
        uom_id = self.env['uom.uom'].search([('name', '=', uom)])
        if uom_id:
            return uom_id.id
        else:
            raise UserError(_('%s Unit Of Measure does not exist in system' % uom))

    def generate_sale_order_line(self, sale_order_id, product, description, quantity, uom, price, tax):
        sale_order_line_id = self.env['sale.order.line'].create({
            'product_id':self._find_product(product),
            'name':description,
            'product_uom_qty':quantity,
            'product_uom':self._find_uom(uom),
            'price_unit':price,
            'order_id':sale_order_id.id
            })
        if tax:
            tax_ids = []
            for record in tax.split(','):
                tax_id = self.env['account.tax'].search([('name', '=', record), ('type_tax_use', '=', 'sale')])
                if not tax_id:
                    raise Warning(_('"%s" Tax not available in your system') % record)
                tax_ids.append(tax_id.id)
                sale_order_line_id.write({'tax_id':([(6, 0, tax_ids)])})

    def generate_sales_order(self, line):
        order_ids = []
        sale_order_id = self.env['sale.order'].search([('origin', '=', line[0])])
        if sale_order_id:
            if sale_order_id.partner_id.name == line[1]:
                if sale_order_id.pricelist_id.name == line[3]:
                    self.generate_sale_order_line(sale_order_id, product=line[5], description=line[6], quantity=line[7], uom=line[8], price=line[9], tax=line[10])
                else:
                    raise UserError(_('Pricelist is different for %s, \n Please define sale' % sale_order_id.id))
            else:
                raise UserError(_('Customer is different for %s, \n Please define sale' % sale_order_id.id))
        else:
            order_id = self.env['sale.order'].create({
                'name':line[0] if self.sequence_option == 'custom' else self.env['ir.sequence'].next_by_code('sale.order') or _('New'),
                'partner_id':self._find_partner(line[1]),
                'date_order':line[2],
                'pricelist_id':self._find_pricelist(line[3]),
                'user_id':self._find_user(line[4]),
                'origin':line[0]
                })
            self.generate_sale_order_line(order_id, product=line[5], description=line[6], quantity=line[7], uom=line[8], price=line[9], tax=line[10])
            order_ids.append(order_id.id)
        if self.quotation_stage_selection == 'confirm':
            self.env['sale.order'].browse(order_ids).action_confirm()
        
    def import_sales_order(self):
        try:
            if self.import_option == 'xls':
                temp_excel_file = tempfile.NamedTemporaryFile(suffix=".xlsx")
                temp_excel_file.write(binascii.a2b_base64(self.file))
                temp_excel_file.seek(0)
                workbook = load_workbook(filename=temp_excel_file.name, data_only=True)
                sheet = workbook.active
                for row_no, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                    if row_no != 1:
                        line = list(map(lambda cell: str(cell) if cell is not None else '', row))
                        try:
                            line[2] = line[2].strftime('%Y-%m-%d') if isinstance(line[2], datetime) else line[2]
                        except Exception:
                            raise UserError(_("Invalid date format in the file."))
                        self.generate_sales_order(line)
            else:
                try:
                    csv_data = base64.b64decode(self.file)
                    try:
                        data_file = io.StringIO(csv_data.decode("utf-8"))
                    except UnicodeDecodeError:
                        data_file = io.StringIO(csv_data.decode("ISO-8859-1"))
                    data_file.seek(0)
                    file_reader = []
                    csv_reader = csv.reader(data_file, delimiter=',')
                    file_reader.extend(csv_reader)
                except Exception as file_error:
                    raise UserError(_("Invalid file: %s") % str(file_error))
                
                for row_no in range(len(file_reader)):
                    line = list(map(str, file_reader[row_no]))
                    if line:
                        if row_no != 0:
                            line[2] = datetime.strptime(line[2], '%d/%m/%y')
                            self.generate_sales_order(line)
        except Exception as e:
            raise UserError(_("Error: %s") % str(e))
