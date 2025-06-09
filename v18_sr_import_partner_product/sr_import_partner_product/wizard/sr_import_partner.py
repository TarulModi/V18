# -*- coding: utf-8 -*-
##############################################################################
#
#    OpenERP, Open Source Management Solution
#    Copyright (C) Sitaram Solutions (<https://sitaramsolutions.in/>).
#
#    For Module Support : info@sitaramsolutions.in  or Skype : contact.hiren1188
#
##############################################################################

import io
import csv
import base64
import logging
import tempfile
import binascii
from openpyxl import load_workbook
from odoo.exceptions import UserError
from odoo import models, fields, api, _


_logger = logging.getLogger(__name__)

class Import_Partner(models.TransientModel):
    _name = 'import.partner'
    _description = 'Import Partner'

    file = fields.Binary('File')
    import_option = fields.Selection([('csv', 'CSV File'), ('xls', 'XLS File')], string='Select', default='csv')
    create_update_option = fields.Selection([('create', 'Create Partner'), ('update', 'Update Partner')], string='Option', default='create')

    def find_state(self, state):
        if not state:
            return False
        state_id = self.env['res.country.state'].search([('name', '=', state)], limit=1)
        if state_id:
            return state_id.id
        else:
            raise UserError(_('%s State does not exist in system' % state))

    def find_country(self, country):
        if not country:
            return False
        country_id = self.env['res.country'].search([('name', '=', country)], limit=1)
        if country_id:
            return country_id.id
        else:
            raise UserError(_('%s country does not exist in system' % country))

    def find_user(self, salesperson):
        if not salesperson:
            return False
        user_id = self.env['res.users'].search([('name', '=', salesperson)], limit=1)
        if user_id:
            return user_id.id
        else:
            raise UserError(_('%s salesperson does not exist in system' % salesperson))

    def create_customer(self, line):
        if self.create_update_option == 'create':
            self.env['res.partner'].create({
                'name': line[0],
                'street': line[1],
                'street2': line[2],
                'city': line[3],
                'state_id': self.find_state(line[4]),
                'zip': line[5].split('.')[0],
                'country_id': self.find_country(line[6]),
                'mobile': line[7].split('.')[0],
                'email': line[8],
                'website': line[9],
                'user_id': self.find_user(line[10]),
            })
        else:
            partner = self.env['res.partner'].search([('name', '=', line[0])], limit=1)
            if partner:
                partner.write({
                    'street': line[1],
                    'street2': line[2],
                    'city': line[3],
                    'state_id': self.find_state(line[4]),
                    'zip': line[5].split('.')[0],
                    'country_id': self.find_country(line[6]),
                    'mobile': line[7].split('.')[0],
                    'email': line[8],
                    'website': line[9],
                    'user_id': self.find_user(line[10]),
                })
            else:
                raise UserError(_('%s Partner does not exist' % line[0]))

    def import_partner(self):
        try:
            if self.import_option == 'xls':
                try:
                    fp = tempfile.NamedTemporaryFile(suffix=".xlsx")
                    fp.write(binascii.a2b_base64(self.file))
                    fp.seek(0)

                    workbook = load_workbook(fp.name, data_only=True)
                    sheet = workbook.active
                    for row_no, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                        if row_no == 1:
                            continue
                        line = [str(cell) if cell is not None else '' for cell in row]
                        self.create_customer(line)
                except Exception as e:
                    _logger.error(f"Error processing XLS file: {str(e)}")
                    raise UserError(str(e))
            else:
                try:
                    csv_data = base64.b64decode(self.file)
                    data_file = io.StringIO(csv_data.decode("utf-8"))
                    csv_reader = csv.reader(data_file, delimiter=',')
                    for i, line in enumerate(csv_reader):
                        if i == 0:
                            continue
                        self.create_customer(line)
                except Exception as e:
                    _logger.error(f"Error processing CSV file: {str(e)}")
                    raise UserError(str(e))
        except Exception as e:
            _logger.error(f"General error during import: {str(e)}")
            raise UserError(str(e))
