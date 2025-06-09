# -*- coding: utf-8 -*-
##############################################################################
#
#    OpenERP, Open Source Management Solution
#    Copyright (C) Sitaram Solutions (<https://sitaramsolutions.in/>).
#
#    For Module Support : info@sitaramsolutions.in  or Skype : contact.hiren1188
#
##############################################################################

import io
import sys
import csv
import xlrd
import base64
import binascii
import tempfile
from datetime import datetime
from openpyxl import load_workbook
from odoo.exceptions import UserError
from odoo import models, fields, api, _


class srImportMultipleBarcode(models.TransientModel):
    _name = 'sr.import.multiple.barcode'
    _description = 'Import Multiple Barcode'

    file = fields.Binary('File')
    import_product_by = fields.Selection([('name', 'Name'), ('code', 'Default Code')], string='Import Product By', default='name')
    import_barcode_for = fields.Selection([('product', 'Products'), ('template', 'Product Template')], string='Import Barcode For', default='product')

    def _import_barcode(self, line):
        domain = []
        if self.import_product_by == 'name':
            if self.import_barcode_for == 'product':
                product_id = self.env['product.product'].search([('name', '=', line[0])])
                if product_id:
                    for barcode in line[1].split(','):
                        self.env['sr.multi.barcode'].create({
                            'name':barcode,
                            'product_id':product_id.id
                            })
                else:
                    raise UserError(_('%s Product Not Found in the system.' %line[0]))
            else:
                product_tmpl_id = self.env['product.template'].search([('name', '=', line[0])])
                if product_tmpl_id:
                    for barcode in line[1].split(','):
                        self.env['sr.multi.barcode'].create({
                            'name':barcode,
                            'product_tmpl_id':product_tmpl_id.id
                            })
                else:
                    raise UserError(_('%s Template Not Found in the system.' %line[0]))
        else:
            if self.import_barcode_for == 'product':
                product_id = self.env['product.product'].search([('default_code', '=', line[0])])
                if product_id:
                    for barcode in line[1].split(','):
                        self.env['sr.multi.barcode'].create({
                            'name':barcode,
                            'product_id':product_id.id
                            })
                else:
                    raise UserError(_('%s Default Code Not Found in the system.' %line[0]))
            else:
                product_tmpl_id = self.env['product.template'].search([('default_code', '=', line[0])])
                if product_tmpl_id:
                    for barcode in line[1].split(','):
                        self.env['sr.multi.barcode'].create({
                            'name':barcode,
                            'product_tmpl_id':product_tmpl_id.id
                            })
                else:
                    raise UserError(_('%s Default Code Not Found in the system.' %line[0]))

    def import_multiple_barcode(self):
        try:
            fp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
            fp.write(binascii.a2b_base64(self.file))
            fp.seek(0)
            fp.close()
            workbook = load_workbook(fp.name, data_only=True)
            sheet = workbook.active
            for row_no, row in enumerate(sheet.iter_rows(values_only=True)):
                if row_no == 0:
                    continue
                line = [str(cell) if cell else '' for cell in row]
                self._import_barcode(line)
        except Exception as e:
            raise UserError(_("Error while importing barcodes: %s") % str(e))
